import openpyxl
import requests
import os

CANVA_API_URL = "https://api.canva.com/v1"

def generate_templates(excel_path, template_url, access_token):
    """Generate customized templates for each participant."""
    # Load participant data from Excel
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    participants = [{"name": row[0].value.strip(), "event": row[1].value.strip()} for row in sheet.iter_rows(min_row=2)]
    workbook.close()

    # Use Canva API to customize templates for each participant
    headers = {
        "Authorization": f"Bearer {access_token}"
    }

    for participant in participants:
        payload = {
            "template_url": template_url,
            "customizations": {
                "name": participant["name"],
                "event": participant["event"]
            }
        }

        response = requests.post(f"{CANVA_API_URL}/templates/customize", json=payload, headers=headers)
        if response.status_code == 200:
            print(f"Template customized for {participant['name']}.")
        else:
            print(f"Failed to customize template for {participant['name']}: {response.text}")

    print("Templates generated. Please download them manually.")

def rename_certificates(excel_path, certificates_folder):
    """Rename certificates based on participant data."""
    # Load participant data from Excel
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    participants = [{"name": row[0].value.strip(), "event": row[1].value.strip()} for row in sheet.iter_rows(min_row=2)]
    workbook.close()

    # Rename certificates
    for index, participant in enumerate(participants, start=1):
        sanitized_name = participant["name"].replace(" ", "_")
        sanitized_event = participant["event"].replace(" ", "_")
        old_filename = os.path.join(certificates_folder, f"{index}.png")  # Assumes files are named numerically
        new_filename = os.path.join(certificates_folder, f"{sanitized_name}_{sanitized_event}.png")
        if os.path.exists(old_filename):
            os.rename(old_filename, new_filename)
            print(f"Renamed: {old_filename} -> {new_filename}")
